import openpyxl
from pathlib import Path
import json
from rich.traceback import install

install(show_locals=True)


class SearchError(Exception):
    pass


def div_id(cell_value: str) -> str:
    code = cell_value.split()[0].strip()
    file: str
    with open(file, "r") as jfile:

        div_list = json.load(jfile)
    for div_data in div_list:
        if div_data["code"] == code:
            return f"{div_data['symbol']}, {div_data['title']}"
    return code


def search_name(
    file_path: Path,
    sheet_name: str,
    table_name: str,
    column_name: str,
    name_to_search: str,
):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    table = [tbl for tbl in sheet.tables if table.name == table_name]
    if table:
        table = table[0]
        column_index = [col.name for col in table.tableColumns].index(column_name) + 1
    else:
        column_index = 5

    results: list = []
    for row in sheet.iter_rows(
        min_col=column_index, max_col=column_index, max_row=3001
    ):
        for cell in row:
            if name_to_search.lower() in str(cell.value).lower():
                reference = f"{cell.column_letter}:{cell.row}"
                action = sheet.cell(row=cell.row, column=cell.column + 2).value
                division = div_id(
                    sheet.cell(row=cell.row, column=cell.column + 12).value
                )
                results.append(
                    {
                        "name": cell.internal_value,
                        "reference": reference,
                        "action": action,
                        "division": division,
                    }
                )
        if not results:
            raise SearchError("No search results found.")
    return results


def main():
    try:
        import yaml

        print(f"\n\n{'--- START ---'}\n\n")

        folder: Path
        data: dict = {}
        for file in folder.iterdir():
            if file.name.startswith("~"):
                continue
            print(file.name)
            try:
                search_results = search_name(
                    file_path=file,
                    sheet_name="TEMPLATE",
                    table_name="template",
                    column_name="NAME",
                    name_to_search="Noecker",
                )
                data[file.name] = search_results
            except SearchError as se:
                print(se)
            except Exception as e:
                print(f"Error: {e}")
            finally:
                print("." * 100, "\n\n")
        data = yaml.safe_dump(data, indent=4)
        print(data)

        print(f"\n\n{'--- FINISH --- '}\n\n")
    except Exception as e:
        print(e)
    except KeyboardInterrupt:
        print("Goodbye!")
